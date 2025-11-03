import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
from datetime import datetime

def count_pdf_pages(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            return len(pdf.pages)
    except Exception as e:
        print(f"Error counting pages: {str(e)}")
        return 0

def extract_sbi_format(page):
    """Extract SBI bank statement format"""
    try:
        # SBI uses clear table structure
        tables = page.extract_tables()
        
        for table in tables:
            if table and len(table) > 1:
                # Check if this is the transaction table
                header = table[0] if table else []
                header_text = ' '.join([str(cell).lower() for cell in header if cell])
                
                # Look for transaction table keywords
                if any(kw in header_text for kw in ['txn date', 'description', 'debit', 'credit', 'balance']):
                    cleaned_table = []
                    for row in table:
                        if row and any(cell and str(cell).strip() for cell in row):
                            cleaned_row = [str(cell).strip() if cell else '' for cell in row]
                            # Skip empty rows
                            if any(cleaned_row):
                                cleaned_table.append(cleaned_row)
                    
                    return cleaned_table
    except Exception as e:
        print(f"SBI extraction error: {str(e)}")
    
    return None

def extract_phonepe_format(page):
    """Extract PhonePe/UPI statement format"""
    try:
        text = page.extract_text()
        if not text:
            return None
        
        lines = text.split('\n')
        transactions = []
        
        # PhonePe format detection
        if 'Transaction Statement' in text or 'phonepe' in text.lower():
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # Date pattern: "Sep 07, 2024" or "Aug 31, 2024"
                date_match = re.match(r'([A-Z][a-z]{2}\s+\d{1,2},\s+\d{4})', line)
                
                if date_match:
                    date = date_match.group(1)
                    time = ''
                    description = ''
                    transaction_type = ''
                    amount = ''
                    transaction_id = ''
                    utr_no = ''
                    
                    # Next line might be time
                    if i + 1 < len(lines):
                        time_line = lines[i + 1].strip()
                        if re.match(r'\d{2}:\d{2}\s+(am|pm)', time_line):
                            time = time_line
                            i += 1
                    
                    # Next lines contain transaction details
                    while i + 1 < len(lines):
                        i += 1
                        detail_line = lines[i].strip()
                        
                        # Check for next date (end of this transaction)
                        if re.match(r'([A-Z][a-z]{2}\s+\d{1,2},\s+\d{4})', detail_line):
                            i -= 1  # Go back one line
                            break
                        
                        # Description (Paid to, Received from)
                        if detail_line.startswith(('Paid to', 'Received from')):
                            description = detail_line
                        
                        # Transaction ID
                        elif detail_line.startswith('Transaction ID'):
                            transaction_id = detail_line.replace('Transaction ID', '').strip()
                        
                        # UTR Number
                        elif detail_line.startswith('UTR No'):
                            utr_no = detail_line.replace('UTR No', '').replace('.', '').strip()
                        
                        # Type (DEBIT/CREDIT) and Amount
                        elif 'DEBIT' in detail_line or 'CREDIT' in detail_line:
                            parts = detail_line.split()
                            for part in parts:
                                if part in ['DEBIT', 'CREDIT']:
                                    transaction_type = part
                                elif part.startswith('₹'):
                                    amount = part.replace('₹', '').replace(',', '')
                        
                        # Check if we hit a separator or page info
                        if 'Page' in detail_line or detail_line.startswith('This is a system'):
                            break
                    
                    # Add transaction if we have minimum data
                    if date and description:
                        transactions.append([
                            date,
                            time,
                            description,
                            transaction_id,
                            utr_no,
                            transaction_type,
                            amount
                        ])
                
                i += 1
            
            if transactions:
                # Add header
                header = ['Date', 'Time', 'Transaction Details', 'Transaction ID', 'UTR No', 'Type', 'Amount']
                return [header] + transactions
    
    except Exception as e:
        print(f"PhonePe extraction error: {str(e)}")
    
    return None

def extract_generic_table(page):
    """Generic table extraction for unknown formats"""
    try:
        tables = page.extract_tables()
        
        for table in tables:
            if table and len(table) > 2:  # At least header + 2 rows
                cleaned_table = []
                for row in table:
                    if row:
                        cleaned_row = [str(cell).strip() if cell else '' for cell in row]
                        if any(cleaned_row):
                            cleaned_table.append(cleaned_row)
                
                if len(cleaned_table) > 2:
                    return cleaned_table
    
    except Exception as e:
        print(f"Generic extraction error: {str(e)}")
    
    return None

def extract_data_from_pdf(pdf_path):
    """Main extraction with format detection"""
    all_tables = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                extracted_table = None
                
                # Try different extraction methods
                # 1. Try SBI format
                extracted_table = extract_sbi_format(page)
                
                # 2. Try PhonePe format
                if not extracted_table:
                    extracted_table = extract_phonepe_format(page)
                
                # 3. Try generic table extraction
                if not extracted_table:
                    extracted_table = extract_generic_table(page)
                
                if extracted_table:
                    all_tables.append({
                        'page': page_num,
                        'data': extracted_table
                    })
    
    except Exception as e:
        print(f"PDF extraction error: {str(e)}")
    
    return all_tables

def normalize_table(table_data):
    """Ensure all rows have same number of columns"""
    if not table_data:
        return []
    
    max_cols = max(len(row) for row in table_data)
    
    normalized = []
    for row in table_data:
        if len(row) < max_cols:
            row = row + [''] * (max_cols - len(row))
        normalized.append(row[:max_cols])
    
    return normalized

def create_excel_from_tables(tables, output_path):
    """Create formatted Excel file"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        normal_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        for table_info in tables:
            table_data = normalize_table(table_info['data'])
            
            if not table_data:
                continue
            
            # Write data
            for row_idx, row_data in enumerate(table_data):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.border = border
                    
                    # First row of each page is header
                    if row_idx == 0:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        cell.font = normal_font
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                current_row += 1
            
            # Add space between pages
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 12), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        return True
    
    except Exception as e:
        print(f"Excel creation error: {str(e)}")
        return False

def convert_pdf_to_excel(pdf_path, output_path):
    """Main conversion function"""
    try:
        pages = count_pdf_pages(pdf_path)
        if pages == 0:
            return False, 0, "Could not read PDF file"
        
        tables = extract_data_from_pdf(pdf_path)
        
        if not tables:
            return False, pages, "No transaction data found in PDF"
        
        success = create_excel_from_tables(tables, output_path)
        
        if success:
            total_rows = sum(len(table['data']) - 1 for table in tables)  # Exclude headers
            return True, pages, f"Converted {pages} page(s) with {total_rows} transaction(s)"
        else:
            return False, pages, "Error creating Excel file"
    
    except Exception as e:
        print(f"Conversion error: {str(e)}")
        return False, 0, f"Error: {str(e)}"
